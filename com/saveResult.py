# -*- coding:utf-8 -*-
# /usr/bin/env python
"""
功能：保存测试结果数据到excel文件，保存的测试结果内容如下：接口，用例，入参，响应数据，预期结果，测试结果，接口响应时间（ms）
保存规则：按测试执行时的时间命名，每执行一次，添加一页sheet，测试结果保存在testResult目录下的test_result.xlsx文件中
Example：
from com.saveResult import SaveResult
sr = SaveResult()
sr.save_result(["接口", "用例", "入参", "响应数据", "预期结果", "测试结果", "接口响应时间"])
"""
import os
from com import readConfig
from com.log import MyLog as Log
from com.tools.operation_excel import OperationExcel as OE
from openpyxl import load_workbook, Workbook
from openpyxl.styles.colors import BLACK
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

pro_dir = readConfig.pro_dir
log = Log(basedir=pro_dir, name="SaveResult.log")
result_path = os.path.join(pro_dir, "testResult")
result_file = os.path.join(result_path, "test_result.xlsx")
# 设置首行单元格式
thin = Side(border_style="thin", color=BLACK)
border = Border(top=thin, left=thin, right=thin, bottom=thin)
title_font = Font(name="黑体", size=12, bold=True, color=BLACK)
title_fill = PatternFill(fill_type='solid', fgColor="4F94CD")
font = Font(name="宋体", size=11, bold=False, color=BLACK)


class SaveResult(object):
    def __init__(self):
        with open(os.path.join(pro_dir, "sheet_name")) as f:
            self.sheet_name = f.read()
            print(self.sheet_name)
        self.__is_first()

    def save_result(self, datas):
        oe = OE(full_name=result_file, sheet=self.sheet_name)
        if not isinstance(datas, list):
            log.error("input datas is not a type list")
            raise TypeError
        oe.write_value(datas=datas)
        oe.close()

    def __is_first(self):
        """
        version:1.0
        1. 判断测试结果文件和测试sheet是否存在，不存在就新创建测试文件和sheet
        2. 设置头行数据的单元格式
        :return:
        """

        if os.path.isfile(result_file):
            wb = load_workbook(filename=result_file)
        else:
            wb = Workbook()
        sheets = wb.sheetnames
        try:
            if self.sheet_name not in sheets:
                wb.create_sheet(self.sheet_name)
                sht = wb[self.sheet_name]
                sht.append([u"接口", u"用例", u"入参", u"响应数据", u"预期响应码", u"预期断言条件", u"接口返回结果", u"测试结果", u"接口响应时间(ms)"])
                for row in sht.rows:
                    for cell in row:
                        cell.border = border
                        cell.font = title_font
                        cell.fill = title_fill
        except Exception as e:
            log.error(str(e))
        finally:
            wb.save(result_file)
            wb.close()


if __name__ == "__main__":
    SaveResult().save_result([123, 45, 789, 56, 23, 12])
    SaveResult().save_result([123, 45, 789, 56, 23, "你好"])
