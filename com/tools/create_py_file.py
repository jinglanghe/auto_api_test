# -*-coding:utf-8 -*-
#! /usr/bin/env python

import os
from openpyxl import Workbook
from openpyxl import load_workbook



class Operate_excel(object):

    def __init__(self, file_path, sheet="Sheet1"):
        self.file_path = file_path
        self.sheet = sheet
        if os.path.isfile(self.file_path):
            self.wb = load_workbook(self.file_path)
        else:
            self.wb = Workbook()

    def read_excel_all_row(self):
        self.sheet_active = self.wb[self.sheet]
        self.sheet_max_row = self.sheet_active.max_row
        self.sheet_max_col = self.sheet_active.max_column


        for row in self.sheet_active.iter_rows(min_row=2, max_row=self.sheet_max_row):

                case_id = str(row[0].value)
                if_run = str(row[1].value)
                goods = str(row[2].value)
                module = str(row[3].value)
                case_name = str(row[4].value)
                case_url = str(row[5].value)
                headers = str(row[6].value)
                method = str(row[7].value)
                params = str(row[8].value)
                params = params.replace('null', 'None')
                params = params.replace('true', 'True')
                params = params.replace('false', 'False')
                respCode = str(row[9].value)
                sqlSen = str(row[10].value)
                api_json_key = str(row[11].value)
                expected_result = str(row[12].value)
                key_sql = str(row[13].value)
                
                py_mid_name = goods + '_' + case_name
                py_file_name = 'test_' +py_mid_name +'.py'

                if if_run == 'yes' :
                    with open('./caselist.txt', 'a',encoding='utf-8') as g:
                        g.write(py_file_name[:-3]+"\n")
                    with open('./testCase/'+py_file_name, 'w',encoding='utf-8') as f:
                        f.write('''
#! /usr/bin/env python
# coding:utf-8

"""
%s
"""

import requests
import json
import time
import datetime
from datetime import timedelta
import unittest
import uuid
import random
from testCase import vesionbook_mall as mall
from com import common
from com import configDB
from com import readConfig
sr = mall.SaveResult()


class Token(unittest.TestCase):
    def setUp(self):
        self._flag = "pass"
        print("this is setup")



    # %s
    def test_%s_caseID_%s(self):
        # 获取token
        gettoken = common.Common.get_token(self)
        _token = str(gettoken)
        
        # 一些可能用到的变量
        _randString = str(uuid.uuid4())[-8:]  
        _randInt = random.randint(0,100)
        today_date = datetime.date.today()
        _today = str(datetime.date.today())
        _yesterday = str(today_date - datetime.timedelta(days=1))
        _8DaysAgo = str(today_date - datetime.timedelta(days=8))
        _lastWeekStart = str(today_date - timedelta(days=today_date.weekday() + 7))
        _lastWeekEnd = str(today_date - timedelta(days=today_date.weekday() + 1))
        
        this_month_start = datetime.datetime(today_date.year, today_date.month, 1)
        last_month_end = (this_month_start - timedelta(days=1)).date()
        last_month_start = datetime.datetime(last_month_end.year, last_month_end.month, 1).date()

        _lastMonthStart = str(last_month_start)
        _lastMonthEnd = str(last_month_end)
        
        _expected_result = %s
        respCode_input = str(%s)
        
        key_sql = %s
        if key_sql == "null" or key_sql ==None:
            pass
        else: 
            db_key = configDB.MysqlHelper()
            key_sql_run_result = db_key.get_all(key_sql)
            _key = str(key_sql_run_result[0][0])
        # 一些可能用到的变量
        
        url = readConfig.ReadBaseConfig().get_http_config('baseurl_user') + "%s"
        if '_key' in url:
            url = url.replace('_key','') + _key
        payload  = %s
        headers = %s
        r = requests.%s(url, data=json.dumps(payload), headers=headers)
        print(r, r.content)
        r_time = float(r.elapsed.microseconds)/1000

        time.sleep(1)
        r_json = r.json()
        print("r_json {}".format(r_json))

        respMessage = r_json["respMessage"]
        print(respMessage)
        r_status = r.status_code
        r_resp_code = str(r_json["respCode"])
        if r_resp_code != respCode_input or r_status != 200:
            self._flag = "fail"
        
        sql = %s
        json_key = %s
        if sql == 'null' or sql == None:
            sql_judge = 'null'
            api_judge = 'null'
            _expected_result = 'null'
        else:
            api_judge = str(r_json%s)
            db_test = configDB.MysqlHelper()
            sql_run_result = db_test.get_all(sql)
            
            if len(sql_run_result) == 0:
                    sql_judge = 'SQL查询为空'
            else:
                sql_judge = str(sql_run_result[0][0])
                    
            if json_key != "null" or json_key != None: 
 
                if sql_judge != api_judge:
                    self._flag = "fail"
                if sql_judge != _expected_result:
                    self._flag = "fail"
            
            if json_key == "null" or json_key == None:
                if sql_judge != _expected_result:
                    self._flag = "fail"    
                    
        
        datas = [url, "%s", str(payload), str(r_json), 
        "resp_code="+ str(respCode_input)+ "\\nstatus=200", sql_judge, api_judge, self._flag, r_time]
        sr.save_result(datas)
        
        self.assertEqual(r_resp_code, respCode_input, msg="返回的状态码不等于"+str(respCode_input)+"\\n"+str(r_json))
        self.assertEqual(_expected_result, sql_judge, msg="Excel填写的预期结果为："+_expected_result+",与SQL查询结果“"+sql_judge+"”不一致。")
        self.assertEqual(api_judge, sql_judge, msg="接口返回结果为："+api_judge+",与SQL查询结果“"+sql_judge+"”不一致。")
        

    if __name__ == "__main__":
        unittest.main()                    
                        '''%(case_name, case_name, goods, case_id, expected_result, respCode, key_sql ,case_url, params, headers, method, sqlSen, api_json_key, api_json_key, case_name))


    def close(self):
        self.wb.save(self.file_path)
        self.wb.close()

if __name__ == "__main__":
    x1 = Operate_excel("../../testCase/testcases.xlsx")
    x1.read_excel_all_row()