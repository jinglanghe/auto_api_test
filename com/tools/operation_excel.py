# -*-coding:utf-8 -*-
#! /usr/bin/env python

import os
from openpyxl import Workbook
from openpyxl import load_workbook
from com.log import MyLog as Log
from com import readConfig


pro_dir = readConfig.pro_dir
log = Log(basedir=pro_dir, name="operation_excel.log")


class OperationExcel(object):
    def __init__(self, full_name, sheet="Sheet1"):
        self.full_name = full_name
        self.sheet = sheet

        if os.path.isfile(self.full_name):
            self.wb = load_workbook(filename=self.full_name)
        else:
            self.wb = Workbook()

    def write_value(self, datas, full_name=None, sheet=None):
        if not isinstance(datas, list):
            log.error("datas not a list type")
            raise TypeError
        if full_name:
            self.full_name = full_name
        if sheet:
            self.sheet = sheet
        sheets = self.wb.sheetnames
        try:
            if self.sheet not in sheets:
                self.wb.create_sheet(self.sheet)
            log.debug(self.wb.sheetnames)
            sht = self.wb[self.sheet]
            sht.append(datas)
        except Exception as e:
            log.error(str(e))

    def close(self):
        self.wb.save(self.full_name)
        self.wb.close()


if __name__ == "__main__":
    data = ["123", "456", "789"]
    xl = OperationExcel("test3.xlsx", "Sheet1")
    xl.write_value(data)
    xl.close()
